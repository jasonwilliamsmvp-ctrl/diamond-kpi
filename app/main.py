import csv
import io
import os
from datetime import date, datetime
from typing import Optional
import json
import calendar

from fastapi import FastAPI, Request, Form, Depends, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from itsdangerous import URLSafeSerializer, BadSignature
from passlib.context import CryptContext
from sqlalchemy import create_engine, String, Integer, Float, Date, DateTime, ForeignKey, Boolean, select, func, text
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, relationship, Session, sessionmaker
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
DATABASE_URL = os.getenv("DATABASE_URL", f"sqlite:///{BASE_DIR}/data/diamond_kpi.db")
# Render provides a PostgreSQL URL. Explicitly select psycopg 3 for SQLAlchemy.
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql+psycopg://", 1)
elif DATABASE_URL.startswith("postgresql://"):
    DATABASE_URL = DATABASE_URL.replace("postgresql://", "postgresql+psycopg://", 1)
SECRET_KEY = os.getenv("SECRET_KEY", "dev-secret-change-me")
COMPANY_NAME = os.getenv("COMPANY_NAME", "晶鑽生醫")
APP_ENV = os.getenv("APP_ENV", "development")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "Admin123!")
SEED_DEMO_DATA = os.getenv("SEED_DEMO_DATA", "true").lower() == "true"
connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(DATABASE_URL, future=True, connect_args=connect_args, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)

class Base(DeclarativeBase):
    pass

class User(Base):
    __tablename__ = "users"
    id: Mapped[int] = mapped_column(primary_key=True)
    username: Mapped[str] = mapped_column(String(80), unique=True, index=True)
    full_name: Mapped[str] = mapped_column(String(100))
    password_hash: Mapped[str] = mapped_column(String(255))
    role: Mapped[str] = mapped_column(String(30), default="sales")
    region: Mapped[str] = mapped_column(String(30), default="全區")
    active: Mapped[bool] = mapped_column(Boolean, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)

class Employee(Base):
    __tablename__ = "employees"
    id: Mapped[int] = mapped_column(primary_key=True)
    employee_no: Mapped[str] = mapped_column(String(30), unique=True, index=True)
    name: Mapped[str] = mapped_column(String(100))
    title: Mapped[str] = mapped_column(String(30), default="專員")
    region: Mapped[str] = mapped_column(String(30), default="北區")
    manager: Mapped[str] = mapped_column(String(100), default="")
    monthly_target: Mapped[float] = mapped_column(Float, default=1000000)
    crm_target: Mapped[float] = mapped_column(Float, default=100)
    visit_target: Mapped[float] = mapped_column(Float, default=40)
    new_clinic_target: Mapped[float] = mapped_column(Float, default=2)
    new_product_target: Mapped[float] = mapped_column(Float, default=1)
    active: Mapped[bool] = mapped_column(Boolean, default=True)

class Product(Base):
    __tablename__ = "products"
    id: Mapped[int] = mapped_column(primary_key=True)
    code: Mapped[str] = mapped_column(String(30), unique=True)
    name: Mapped[str] = mapped_column(String(100))
    category: Mapped[str] = mapped_column(String(60), default="注射產品")
    unit: Mapped[str] = mapped_column(String(20), default="盒")
    unit_price: Mapped[float] = mapped_column(Float, default=0)
    gross_margin: Mapped[float] = mapped_column(Float, default=0.7)
    monthly_target_qty: Mapped[float] = mapped_column(Float, default=0)
    active: Mapped[bool] = mapped_column(Boolean, default=True)

class Clinic(Base):
    __tablename__ = "clinics"
    id: Mapped[int] = mapped_column(primary_key=True)
    code: Mapped[str] = mapped_column(String(30), unique=True)
    name: Mapped[str] = mapped_column(String(150))
    region: Mapped[str] = mapped_column(String(30), default="北區")
    city: Mapped[str] = mapped_column(String(50), default="台北市")
    owner_employee_id: Mapped[Optional[int]] = mapped_column(ForeignKey("employees.id"), nullable=True)
    status: Mapped[str] = mapped_column(String(30), default="有效客戶")
    last_order_date: Mapped[Optional[date]] = mapped_column(Date, nullable=True)
    owner: Mapped[Optional[Employee]] = relationship()

class Sale(Base):
    __tablename__ = "sales"
    id: Mapped[int] = mapped_column(primary_key=True)
    sale_date: Mapped[date] = mapped_column(Date, default=date.today, index=True)
    employee_id: Mapped[int] = mapped_column(ForeignKey("employees.id"))
    product_id: Mapped[int] = mapped_column(ForeignKey("products.id"))
    clinic_id: Mapped[int] = mapped_column(ForeignKey("clinics.id"))
    quantity: Mapped[float] = mapped_column(Float, default=1)
    amount: Mapped[float] = mapped_column(Float, default=0)
    gross_profit: Mapped[float] = mapped_column(Float, default=0)
    status: Mapped[str] = mapped_column(String(30), default="已認列")
    note: Mapped[str] = mapped_column(String(255), default="")
    employee: Mapped[Employee] = relationship()
    product: Mapped[Product] = relationship()
    clinic: Mapped[Clinic] = relationship()

class Activity(Base):
    __tablename__ = "activities"
    id: Mapped[int] = mapped_column(primary_key=True)
    activity_date: Mapped[date] = mapped_column(Date, default=date.today)
    employee_id: Mapped[int] = mapped_column(ForeignKey("employees.id"))
    clinic_id: Mapped[int] = mapped_column(ForeignKey("clinics.id"))
    stage: Mapped[str] = mapped_column(String(30), default="拜訪")
    outcome: Mapped[str] = mapped_column(String(100), default="")
    next_action_date: Mapped[Optional[date]] = mapped_column(Date, nullable=True)
    employee: Mapped[Employee] = relationship()
    clinic: Mapped[Clinic] = relationship()

class AuditLog(Base):
    __tablename__ = "audit_logs"
    id: Mapped[int] = mapped_column(primary_key=True)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    username: Mapped[str] = mapped_column(String(80))
    action: Mapped[str] = mapped_column(String(100))
    entity: Mapped[str] = mapped_column(String(50))
    detail: Mapped[str] = mapped_column(String(500), default="")

pwd = CryptContext(schemes=["bcrypt"], deprecated="auto")
signer = URLSafeSerializer(SECRET_KEY, salt="diamond-session")
app = FastAPI(title="Diamond KPI Enterprise")
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "app", "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "app", "templates"))


def db_session():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def current_user(request: Request, db: Session = Depends(db_session)) -> User:
    token = request.cookies.get("diamond_session")
    if not token:
        raise HTTPException(401)
    try:
        data = signer.loads(token)
    except BadSignature:
        raise HTTPException(401)
    user = db.get(User, data.get("uid"))
    if not user or not user.active:
        raise HTTPException(401)
    return user


def audit(db: Session, user: User, action: str, entity: str, detail: str = ""):
    db.add(AuditLog(username=user.username, action=action, entity=entity, detail=detail[:500]))
    db.commit()


def authorize(user: User, *roles: str):
    if user.role not in roles:
        raise HTTPException(403, "權限不足")

@app.exception_handler(401)
async def unauthorized(request: Request, exc):
    return RedirectResponse("/login", status_code=303)

@app.on_event("startup")
def startup():
    Base.metadata.create_all(engine)
    # Lightweight schema migration for existing Render/PostgreSQL databases.
    # create_all() does not add new columns to an existing table.
    with engine.begin() as conn:
        if engine.dialect.name == "postgresql":
            conn.execute(text("ALTER TABLE employees ADD COLUMN IF NOT EXISTS crm_target DOUBLE PRECISION DEFAULT 100"))
            conn.execute(text("ALTER TABLE employees ADD COLUMN IF NOT EXISTS visit_target DOUBLE PRECISION DEFAULT 40"))
            conn.execute(text("ALTER TABLE employees ADD COLUMN IF NOT EXISTS new_clinic_target DOUBLE PRECISION DEFAULT 2"))
            conn.execute(text("ALTER TABLE employees ADD COLUMN IF NOT EXISTS new_product_target DOUBLE PRECISION DEFAULT 1"))
        elif engine.dialect.name == "sqlite":
            cols = {r[1] for r in conn.execute(text("PRAGMA table_info(employees)"))}
            for col, default in [("crm_target",100),("visit_target",40),("new_clinic_target",2),("new_product_target",1)]:
                if col not in cols:
                    conn.execute(text(f"ALTER TABLE employees ADD COLUMN {col} FLOAT DEFAULT {default}"))
    db = SessionLocal()
    try:
        if not db.scalar(select(func.count(User.id))):
            users = [
                User(username=ADMIN_USERNAME, full_name="系統管理員", password_hash=pwd.hash(ADMIN_PASSWORD), role="admin", region="全區")
            ]
            if SEED_DEMO_DATA:
                users.extend([
                    User(username="ceo", full_name="總經理", password_hash=pwd.hash("Ceo123!"), role="executive", region="全區"),
                    User(username="manager", full_name="北區經理", password_hash=pwd.hash("Manager123!"), role="manager", region="北區"),
                    User(username="sales", full_name="業務示範", password_hash=pwd.hash("Sales123!"), role="sales", region="北區"),
                ])
            db.add_all(users)
        if not db.scalar(select(func.count(Employee.id))):
            emps = [
                Employee(employee_no="E001", name="陳協理", title="協理", region="全區", monthly_target=50000000),
                Employee(employee_no="E101", name="王區經理", title="區域經理", region="北區", monthly_target=15000000),
                Employee(employee_no="E201", name="林襄理", title="襄理", region="北區", manager="王區經理", monthly_target=2500000),
                Employee(employee_no="E202", name="張主任", title="主任", region="中區", manager="陳協理", monthly_target=1800000),
                Employee(employee_no="E203", name="李專員", title="專員", region="南區", manager="陳協理", monthly_target=1000000),
            ]
            db.add_all(emps); db.flush()
            prods = [
                Product(code="MET", name="METEORA", unit="盒", unit_price=120000, gross_margin=.72, monthly_target_qty=20),
                Product(code="NEO", name="NeoFilera", unit="瓶", unit_price=80000, gross_margin=.75, monthly_target_qty=30),
                Product(code="NVB", name="NovaBright", unit="台", unit_price=0, gross_margin=.70, monthly_target_qty=0),
                Product(code="RON", name="Ronkylä", unit="盒", unit_price=60000, gross_margin=.68, monthly_target_qty=50),
                Product(code="PK", name="Pico-K", category="儀器", unit="台", unit_price=1500000, gross_margin=.55, monthly_target_qty=2),
                Product(code="PT", name="探頭系列", category="耗材", unit="支", unit_price=25000, gross_margin=.65, monthly_target_qty=80),
            ]
            db.add_all(prods); db.flush()
            clinics = [Clinic(code=f"C{i:03d}", name=n, region=r, city=c, owner_employee_id=emps[min(i-1,4)].id) for i,(n,r,c) in enumerate([
                ("晶采醫美診所","北區","台北市"),("澄美醫美診所","北區","新北市"),("璞研診所","中區","台中市"),("雅緻醫美診所","南區","高雄市"),("悅容診所","南區","台南市")],1)]
            db.add_all(clinics); db.flush()
            today=date.today()
            demo_sales=[
                (emps[2],prods[0],clinics[0],12,1440000),(emps[2],prods[1],clinics[1],18,1440000),
                (emps[3],prods[2],clinics[2],24,1440000),(emps[4],prods[4],clinics[3],38,950000),
                (emps[1],prods[3],clinics[0],1,1500000),(emps[0],prods[0],clinics[4],8,960000),
            ]
            for e,p,c,q,a in demo_sales:
                db.add(Sale(sale_date=today, employee_id=e.id, product_id=p.id, clinic_id=c.id, quantity=q, amount=a, gross_profit=a*p.gross_margin))
            stages=["拜訪","拜訪","提案","報價","成交","回購","拜訪","提案"]
            for i,st in enumerate(stages):
                db.add(Activity(activity_date=today, employee_id=emps[i%len(emps)].id, clinic_id=clinics[i%len(clinics)].id, stage=st, outcome="示範資料"))
        # Ensure NovaBright is present even on an existing Render/PostgreSQL database.
        if not db.scalar(select(Product).where(Product.code == "NVB")):
            db.add(Product(code="NVB", name="NovaBright", category="設備", unit="台", unit_price=0, gross_margin=.70, monthly_target_qty=0, active=True))
        db.commit()
    finally:
        db.close()


@app.get("/health")
def health():
    return {"status": "ok", "service": "diamond-kpi"}

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "company": COMPANY_NAME})

@app.post("/login")
def login(username: str = Form(...), password: str = Form(...), db: Session = Depends(db_session)):
    user = db.scalar(select(User).where(User.username == username))
    if not user or not pwd.verify(password, user.password_hash):
        return RedirectResponse("/login?error=1", status_code=303)
    response = RedirectResponse("/", status_code=303)
    response.set_cookie(
        "diamond_session",
        signer.dumps({"uid": user.id}),
        httponly=True,
        secure=(APP_ENV == "production"),
        samesite="lax",
        max_age=28800,
    )
    return response

@app.get("/logout")
def logout():
    r=RedirectResponse("/login",303); r.delete_cookie("diamond_session"); return r


def _parse_month(month: Optional[str]):
    try:
        if month:
            y, m = [int(x) for x in month.split("-")]
            return date(y, m, 1)
    except Exception:
        pass
    return date.today().replace(day=1)


def _month_end(month_start: date):
    return date(month_start.year, month_start.month, calendar.monthrange(month_start.year, month_start.month)[1])


def _shift_month(month_start: date, delta: int):
    idx = month_start.year * 12 + (month_start.month - 1) + delta
    return date(idx // 12, idx % 12 + 1, 1)


def _pct(actual, target):
    return (actual / target * 100) if target else 0


def _light(rate: float, hard_red: bool = False):
    if hard_red or rate < 80:
        return "red"
    if rate < 100:
        return "yellow"
    return "green"


def _title_target(title: str):
    return {
        "專員": 1_000_000,
        "主任": 1_800_000,
        "襄理": 2_500_000,
        "區域經理": 15_000_000,
        "協理": 50_000_000,
    }.get(title, 1_000_000)


def _promotion_text(title: str):
    order=["專員","主任","襄理","區域經理","協理"]
    if title not in order or title == "協理":
        return "維持職級 / 高績效獎勵"
    return f"可晉升 {order[order.index(title)+1]}"


def _team_member_ids(db: Session, e: Employee):
    """Return the employee IDs that belong to a manager's KPI team.

    Prefer the explicit manager hierarchy. If the hierarchy is incomplete,
    regional managers fall back to their region and a top-level 協理 falls
    back to all active employees. The manager's own sales are included.
    """
    if e.title not in ("區域經理", "協理"):
        return [e.id]

    active=list(db.scalars(select(Employee).where(Employee.active==True)))
    by_manager={}
    for emp in active:
        key=(emp.manager or "").strip()
        if key:
            by_manager.setdefault(key, []).append(emp)

    ids={e.id}
    queue=[e.name]
    seen_names=set()
    while queue:
        manager_name=queue.pop(0)
        if manager_name in seen_names:
            continue
        seen_names.add(manager_name)
        for child in by_manager.get(manager_name, []):
            if child.id not in ids:
                ids.add(child.id)
                queue.append(child.name)

    # Fallback for legacy data where reporting lines were not completely maintained.
    if len(ids) == 1:
        if e.title == "區域經理":
            ids.update(emp.id for emp in active if emp.region == e.region)
        elif e.title == "協理":
            ids.update(emp.id for emp in active)
    return list(ids)


def _employee_month_rate(db: Session, e: Employee, month_start: date):
    month_end=_month_end(month_start)
    member_ids=_team_member_ids(db,e)
    amt=db.scalar(select(func.coalesce(func.sum(Sale.amount),0)).where(
        Sale.employee_id.in_(member_ids), Sale.sale_date>=month_start, Sale.sale_date<=month_end
    )) or 0
    target=e.monthly_target or _title_target(e.title)
    return float(amt), _pct(float(amt), float(target))


def _scope_team_target(emps):
    """Pick one non-overlapping target layer for the current dashboard scope.

    This prevents double counting the same organization by adding 協理、區經理、
    襄理、主任、專員 thresholds together. We use the highest managerial layer
    that exists in the current scope, and only fall back to individual targets
    when no team manager exists.
    """
    for title in ("協理", "區域經理"):
        managers=[e for e in emps if e.title == title]
        if managers:
            return sum(float(e.monthly_target or _title_target(e.title)) for e in managers)
    return sum(float(e.monthly_target or _title_target(e.title)) for e in emps)


def _consecutive_status(db: Session, e: Employee, month_start: date):
    achieved=0; missed=0
    for i in range(0, 6):
        m=_shift_month(month_start,-i)
        _, rate=_employee_month_rate(db,e,m)
        if i == 0:
            current_rate=rate
        if rate >= 100:
            if missed == 0: achieved += 1
            else: break
        else:
            if achieved == 0: missed += 1
            else: break
    return achieved, missed, current_rate


def kpi_context(db: Session, user: User, month: Optional[str] = None):
    month_start=_parse_month(month)
    month_end=_month_end(month_start)
    month_key=month_start.strftime("%Y-%m")
    sales_query=select(Sale).where(Sale.sale_date>=month_start, Sale.sale_date<=month_end)
    if user.role in ("manager","sales"):
        sales_query=sales_query.join(Employee).where(Employee.region==user.region)
    sales=list(db.scalars(sales_query))

    emps=list(db.scalars(select(Employee).where(Employee.active==True)))
    if user.role in ("manager","sales"):
        emps=[e for e in emps if e.region==user.region]

    revenue=sum(s.amount for s in sales)
    gp=sum(s.gross_profit for s in sales)
    # Team target uses only one organizational layer to avoid double counting
    # managerial team thresholds together with subordinate individual thresholds.
    target=_scope_team_target(emps)
    rate=_pct(revenue,target)
    avg_value=revenue/len(emps) if emps else 0
    avg_target=(target/len(emps)) if emps else 0

    activities_q=select(Activity).where(Activity.activity_date>=month_start, Activity.activity_date<=month_end)
    if user.role in ("manager","sales"):
        activities_q=activities_q.join(Employee).where(Employee.region==user.region)
    activities=list(db.scalars(activities_q))
    visits=sum(1 for a in activities if a.stage=="拜訪")
    completed=sum(1 for a in activities if (a.outcome or '').strip() and a.next_action_date is not None)
    crm_complete=_pct(completed,len(activities)) if activities else 0

    # New effective ordering clinics = clinics whose first recorded sale falls inside selected month.
    new_ordering=0
    relevant_clinic_ids={s.clinic_id for s in sales}
    for cid in relevant_clinic_ids:
        first=db.scalar(select(func.min(Sale.sale_date)).where(Sale.clinic_id==cid))
        if first and month_start <= first <= month_end:
            new_ordering += 1

    # New product introductions = first-ever clinic-product pair occurring this month.
    intro_pairs=set()
    for s in sales:
        first=db.scalar(select(func.min(Sale.sale_date)).where(Sale.clinic_id==s.clinic_id, Sale.product_id==s.product_id))
        if first and month_start <= first <= month_end:
            intro_pairs.add((s.clinic_id,s.product_id))
    new_product_intro=len({cid for cid,_ in intro_pairs})

    sales_kpis=[
        {"name":"團隊業績","target":target,"actual":revenue,"unit":"元","rate":rate,"light":_light(rate),"note":f"距目標差 {max(target-revenue,0):,.0f} 元" if rate<100 else "已達成團隊目標"},
        {"name":"團隊人均產值","target":avg_target,"actual":avg_value,"unit":"元／人","rate":_pct(avg_value,avg_target),"light":_light(_pct(avg_value,avg_target)),"note":f"共 {len(emps)} 位有效人員"},
        {"name":"管理幅度","target":8,"actual":len([e for e in emps if e.title in ['專員','主任','襄理']]),"unit":"人","rate":min(_pct(len([e for e in emps if e.title in ['專員','主任','襄理']]),8),100),"light":"green" if len([e for e in emps if e.title in ['專員','主任','襄理']])<=8 else "yellow","note":"建議每位主管管理幅度 ≤ 8 人"},
        {"name":"高毛利回報","target":70,"actual":(gp/revenue*100 if revenue else 0),"unit":"%","rate":_pct((gp/revenue*100 if revenue else 0),70),"light":_light(_pct((gp/revenue*100 if revenue else 0),70)),"note":"以整體毛利率 70% 作為管理基準"},
    ]
    crm_kpis=[
        {"name":"新增有效下單診所","target":2,"actual":new_ordering,"unit":"家","rate":_pct(new_ordering,2),"light":_light(_pct(new_ordering,2)),"note":"目標 ≥ 2 家／月"},
        {"name":"客戶拜訪數","target":40,"actual":visits,"unit":"家","rate":_pct(visits,40),"light":_light(_pct(visits,40)),"note":"目標 ≥ 40 家／月"},
        {"name":"新品導入數","target":1,"actual":new_product_intro,"unit":"家","rate":_pct(new_product_intro,1),"light":_light(_pct(new_product_intro,1)),"note":"首次 clinic-product 組合視為新品導入"},
        {"name":"CRM 完整度","target":100,"actual":crm_complete,"unit":"%","rate":crm_complete,"light":_light(crm_complete),"note":"結果 + 下一步日期皆填寫才算完整"},
    ]

    by_emp=[]
    for e in emps:
        target_e=e.monthly_target or _title_target(e.title)
        if e.title in ("協理", "區域經理"):
            team_ids=set(_team_member_ids(db,e))
            emp_sales=[s for s in sales if s.employee_id in team_ids]
        else:
            emp_sales=[s for s in sales if s.employee_id==e.id]
        amt=sum(s.amount for s in emp_sales)
        rev_rate=_pct(amt,target_e)
        emp_acts=[a for a in activities if a.employee_id==e.id]
        emp_crm=_pct(sum(1 for a in emp_acts if (a.outcome or '').strip() and a.next_action_date is not None),len(emp_acts)) if emp_acts else 0
        achieved,missed,_=_consecutive_status(db,e,month_start)
        if achieved>=3:
            status="green"; status_text="綠燈"; action=_promotion_text(e.title)
        elif missed>=3:
            status="red"; status_text="紅燈"; action="降職／調整職位"
        elif missed>=2:
            status="red"; status_text="紅燈"; action="黃牌輔導；次月未改善進入淘汰評估"
        elif rev_rate<80 or emp_crm<80:
            status="red"; status_text="紅燈"; action="立即輔導改善"
        elif rev_rate<100 or emp_crm<100:
            status="yellow"; status_text="黃燈"; action="持續觀察"
        else:
            status="green"; status_text="綠燈"; action="達標"
        by_emp.append({"id":e.id,"name":e.name,"title":e.title,"region":e.region,"amount":amt,"target":target_e,"revenue_rate":rev_rate,"crm_rate":emp_crm,"crm_target":(e.crm_target if e.crm_target is not None else 100),"status":status,"status_text":status_text,"action":action,"achieved_months":achieved,"missed_months":missed})
    by_emp.sort(key=lambda x:(x["status"]!="red", x["revenue_rate"]))

    warning_count=sum(1 for r in by_emp if r["status"]=="red")
    yellow_count=sum(1 for r in by_emp if r["status"]=="yellow")
    prev_month=_shift_month(month_start,-1).strftime("%Y-%m")
    next_month=_shift_month(month_start,1).strftime("%Y-%m")
    return {
        "month_start":month_start,"month_key":month_key,"prev_month":prev_month,"next_month":next_month,
        "revenue":revenue,"target":target,"rate":rate,"gp":gp,"margin":gp/revenue*100 if revenue else 0,
        "avg_value":avg_value,"avg_target":avg_target,"crm_complete":crm_complete,"new_ordering":new_ordering,"visits":visits,
        "sales_kpis":sales_kpis,"crm_kpis":crm_kpis,"by_emp":by_emp,"warning_count":warning_count,"yellow_count":yellow_count,
    }


@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, month: Optional[str]=None, db: Session=Depends(db_session), user: User=Depends(current_user)):
    ctx=kpi_context(db,user,month)
    return templates.TemplateResponse("dashboard.html", {"request":request,"user":user,"company":COMPANY_NAME,**ctx})


@app.get("/api/dashboard")
def api_dashboard(month: Optional[str]=None, db: Session=Depends(db_session), user: User=Depends(current_user)):
    return kpi_context(db,user,month)


@app.get("/export/kpi.csv")
def export_kpi(month: Optional[str]=None, db:Session=Depends(db_session), user:User=Depends(current_user)):
    ctx=kpi_context(db,user,month)
    out=io.StringIO(); w=csv.writer(out)
    w.writerow(["月份","員工","職級","區域","業績目標","實際業績","業績達成率","CRM完整度","燈號","連續達標月","連續未達標月","管理建議"])
    for r in ctx["by_emp"]:
        w.writerow([ctx["month_key"],r["name"],r["title"],r["region"],r["target"],r["amount"],round(r["revenue_rate"],1),round(r["crm_rate"],1),r["status_text"],r["achieved_months"],r["missed_months"],r["action"]])
    data=out.getvalue().encode("utf-8-sig")
    return StreamingResponse(io.BytesIO(data),media_type="text/csv",headers={"Content-Disposition":f"attachment; filename=diamond_kpi_{ctx['month_key']}.csv"})

@app.get("/employees", response_class=HTMLResponse)
def employees_page(request:Request, db:Session=Depends(db_session), user:User=Depends(current_user)):
    rows=list(db.scalars(select(Employee).order_by(Employee.region,Employee.title)))
    return templates.TemplateResponse("employees.html",{"request":request,"user":user,"company":COMPANY_NAME,"rows":rows})

@app.post("/employees")
def employee_add(employee_no:str=Form(...),name:str=Form(...),title:str=Form(...),region:str=Form(...),manager:str=Form(""),monthly_target:float=Form(...),crm_target:float=Form(100),visit_target:float=Form(40),new_clinic_target:float=Form(2),new_product_target:float=Form(1),db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin","executive")
    db.add(Employee(employee_no=employee_no,name=name,title=title,region=region,manager=manager,monthly_target=monthly_target,crm_target=crm_target,visit_target=visit_target,new_clinic_target=new_clinic_target,new_product_target=new_product_target)); db.commit(); audit(db,user,"新增","員工",f"{employee_no} {name}")
    return RedirectResponse("/employees",303)

@app.post("/employees/{eid}/kpi")
def employee_kpi_update(eid:int, monthly_target:float=Form(...), crm_target:float=Form(...), visit_target:float=Form(...), new_clinic_target:float=Form(...), new_product_target:float=Form(...), db:Session=Depends(db_session), user:User=Depends(current_user)):
    authorize(user,"admin","executive")
    e=db.get(Employee,eid)
    if not e: raise HTTPException(404)
    e.monthly_target=max(monthly_target,0); e.crm_target=min(max(crm_target,0),100); e.visit_target=max(visit_target,0); e.new_clinic_target=max(new_clinic_target,0); e.new_product_target=max(new_product_target,0)
    db.commit(); audit(db,user,"更新KPI","員工",f"{e.employee_no} {e.name} 業績={e.monthly_target}, CRM={e.crm_target}%, 拜訪={e.visit_target}, 新診所={e.new_clinic_target}, 新品={e.new_product_target}")
    return RedirectResponse("/employees?saved=1",303)

@app.post("/employees/{eid}/delete")
def employee_delete(eid:int,db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin")
    e=db.get(Employee,eid)
    if e: e.active=False; db.commit(); audit(db,user,"停用","員工",e.name)
    return RedirectResponse("/employees",303)

@app.get("/sales", response_class=HTMLResponse)
def sales_page(request:Request,db:Session=Depends(db_session),user:User=Depends(current_user)):
    rows=list(db.scalars(select(Sale).order_by(Sale.sale_date.desc(),Sale.id.desc()).limit(300)))
    emps=list(db.scalars(select(Employee).where(Employee.active==True))); products=list(db.scalars(select(Product).where(Product.active==True))); clinics=list(db.scalars(select(Clinic)))
    return templates.TemplateResponse("sales.html",{"request":request,"user":user,"company":COMPANY_NAME,"rows":rows,"employees":emps,"products":products,"clinics":clinics})

@app.post("/sales")
def sale_add(sale_date:date=Form(...),employee_id:int=Form(...),product_id:int=Form(...),clinic_id:int=Form(...),quantity:float=Form(...),amount:float=Form(...),note:str=Form(""),db:Session=Depends(db_session),user:User=Depends(current_user)):
    p=db.get(Product,product_id); gp=amount*(p.gross_margin if p else 0)
    db.add(Sale(sale_date=sale_date,employee_id=employee_id,product_id=product_id,clinic_id=clinic_id,quantity=quantity,amount=amount,gross_profit=gp,note=note));
    c=db.get(Clinic,clinic_id)
    if c: c.last_order_date=sale_date
    db.commit(); audit(db,user,"新增","業績",f"金額 {amount:,.0f}")
    return RedirectResponse("/sales",303)

@app.get("/sales/{sid}/edit", response_class=HTMLResponse)
def sale_edit_page(sid:int, request:Request, db:Session=Depends(db_session), user:User=Depends(current_user)):
    authorize(user,"admin","executive","manager")
    sale=db.get(Sale,sid)
    if not sale:
        raise HTTPException(404,"找不到業績資料")
    emps=list(db.scalars(select(Employee).where(Employee.active==True).order_by(Employee.region,Employee.name)))
    products=list(db.scalars(select(Product).where(Product.active==True).order_by(Product.name)))
    clinics=list(db.scalars(select(Clinic).order_by(Clinic.region,Clinic.name)))
    return templates.TemplateResponse("sales_edit.html",{
        "request":request,"user":user,"company":COMPANY_NAME,"sale":sale,
        "employees":emps,"products":products,"clinics":clinics
    })

@app.post("/sales/{sid}/edit")
def sale_edit(sid:int,sale_date:date=Form(...),employee_id:int=Form(...),product_id:int=Form(...),clinic_id:int=Form(...),quantity:float=Form(...),amount:float=Form(...),status:str=Form("已認列"),note:str=Form(""),db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin","executive","manager")
    sale=db.get(Sale,sid)
    if not sale:
        raise HTTPException(404,"找不到業績資料")
    product=db.get(Product,product_id)
    sale.sale_date=sale_date
    sale.employee_id=employee_id
    sale.product_id=product_id
    sale.clinic_id=clinic_id
    sale.quantity=quantity
    sale.amount=amount
    sale.gross_profit=amount*(product.gross_margin if product else 0)
    sale.status=status
    sale.note=note
    clinic=db.get(Clinic,clinic_id)
    if clinic:
        clinic.last_order_date=sale_date
    db.commit()
    audit(db,user,"修改","業績",f"#{sid} 金額 {amount:,.0f}")
    return RedirectResponse("/sales",303)

@app.post("/sales/{sid}/delete")
def sale_delete(sid:int,db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin","executive","manager")
    s=db.get(Sale,sid)
    if s: db.delete(s); db.commit(); audit(db,user,"刪除","業績",str(sid))
    return RedirectResponse("/sales",303)

@app.get("/products", response_class=HTMLResponse)
def products_page(request:Request,db:Session=Depends(db_session),user:User=Depends(current_user)):
    rows=list(db.scalars(select(Product).order_by(Product.name)))
    return templates.TemplateResponse("products.html",{"request":request,"user":user,"company":COMPANY_NAME,"rows":rows})

@app.post("/products")
def product_add(code:str=Form(...),name:str=Form(...),category:str=Form(...),unit:str=Form(...),unit_price:float=Form(...),gross_margin:float=Form(...),monthly_target_qty:float=Form(...),db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin","executive")
    db.add(Product(code=code,name=name,category=category,unit=unit,unit_price=unit_price,gross_margin=gross_margin/100,monthly_target_qty=monthly_target_qty)); db.commit(); audit(db,user,"新增","產品",name)
    return RedirectResponse("/products",303)

@app.get("/clinics",response_class=HTMLResponse)
def clinics_page(request:Request,db:Session=Depends(db_session),user:User=Depends(current_user)):
    rows=list(db.scalars(select(Clinic).order_by(Clinic.region,Clinic.name))); emps=list(db.scalars(select(Employee).where(Employee.active==True)))
    return templates.TemplateResponse("clinics.html",{"request":request,"user":user,"company":COMPANY_NAME,"rows":rows,"employees":emps})

@app.post("/clinics")
def clinic_add(code:str=Form(...),name:str=Form(...),region:str=Form(...),city:str=Form(...),owner_employee_id:int=Form(...),status:str=Form(...),db:Session=Depends(db_session),user:User=Depends(current_user)):
    db.add(Clinic(code=code,name=name,region=region,city=city,owner_employee_id=owner_employee_id,status=status)); db.commit(); audit(db,user,"新增","診所",name)
    return RedirectResponse("/clinics",303)

@app.get("/activities",response_class=HTMLResponse)
def activities_page(request:Request,db:Session=Depends(db_session),user:User=Depends(current_user)):
    rows=list(db.scalars(select(Activity).order_by(Activity.activity_date.desc(),Activity.id.desc()).limit(300))); emps=list(db.scalars(select(Employee).where(Employee.active==True))); clinics=list(db.scalars(select(Clinic)))
    return templates.TemplateResponse("activities.html",{"request":request,"user":user,"company":COMPANY_NAME,"rows":rows,"employees":emps,"clinics":clinics})

@app.post("/activities")
def activity_add(activity_date:date=Form(...),employee_id:int=Form(...),clinic_id:int=Form(...),stage:str=Form(...),outcome:str=Form(""),next_action_date:Optional[date]=Form(None),db:Session=Depends(db_session),user:User=Depends(current_user)):
    db.add(Activity(activity_date=activity_date,employee_id=employee_id,clinic_id=clinic_id,stage=stage,outcome=outcome,next_action_date=next_action_date)); db.commit(); audit(db,user,"新增","CRM活動",stage)
    return RedirectResponse("/activities",303)

@app.get("/users",response_class=HTMLResponse)
def users_page(request:Request,db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin"); rows=list(db.scalars(select(User).order_by(User.username)))
    return templates.TemplateResponse("users.html",{"request":request,"user":user,"company":COMPANY_NAME,"rows":rows})

@app.post("/users")
def user_add(username:str=Form(...),full_name:str=Form(...),password:str=Form(...),role:str=Form(...),region:str=Form(...),db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin"); db.add(User(username=username,full_name=full_name,password_hash=pwd.hash(password),role=role,region=region)); db.commit(); audit(db,user,"新增","使用者",username)
    return RedirectResponse("/users",303)

@app.get("/audit",response_class=HTMLResponse)
def audit_page(request:Request,db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin","executive"); rows=list(db.scalars(select(AuditLog).order_by(AuditLog.created_at.desc()).limit(500)))
    return templates.TemplateResponse("audit.html",{"request":request,"user":user,"company":COMPANY_NAME,"rows":rows})

@app.get("/export/sales.csv")
def export_sales(db:Session=Depends(db_session),user:User=Depends(current_user)):
    out=io.StringIO(); w=csv.writer(out); w.writerow(["日期","員工編號","員工","產品代碼","產品","診所代碼","診所","數量","金額","毛利","狀態","備註"])
    for s in db.scalars(select(Sale).order_by(Sale.sale_date)):
        w.writerow([s.sale_date,s.employee.employee_no,s.employee.name,s.product.code,s.product.name,s.clinic.code,s.clinic.name,s.quantity,s.amount,s.gross_profit,s.status,s.note])
    data=out.getvalue().encode("utf-8-sig")
    return StreamingResponse(io.BytesIO(data),media_type="text/csv",headers={"Content-Disposition":"attachment; filename=diamond_sales.csv"})

def _read_tabular_upload(filename: str, raw: bytes):
    name=(filename or "").lower()
    if name.endswith(".xlsx"):
        wb=load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws=wb.active
        values=list(ws.iter_rows(values_only=True))
        if not values: return []
        headers=[str(x).strip() if x is not None else "" for x in values[0]]
        return [dict(zip(headers,row)) for row in values[1:] if any(v not in (None,"") for v in row)]
    text=raw.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))

def _norm(v):
    return "" if v is None else str(v).strip()

def _sales_validate(rows, db):
    checked=[]
    required=["日期","員工編號","產品代碼","診所代碼","數量","金額"]
    for idx,row in enumerate(rows, start=2):
        errors=[]
        for k in required:
            if _norm(row.get(k))=="": errors.append(f"缺少{k}")
        emp=db.scalar(select(Employee).where(Employee.employee_no==_norm(row.get("員工編號")))) if _norm(row.get("員工編號")) else None
        prod=db.scalar(select(Product).where(Product.code==_norm(row.get("產品代碼")))) if _norm(row.get("產品代碼")) else None
        clinic=db.scalar(select(Clinic).where(Clinic.code==_norm(row.get("診所代碼")))) if _norm(row.get("診所代碼")) else None
        if _norm(row.get("員工編號")) and not emp: errors.append("員工編號不存在")
        if _norm(row.get("產品代碼")) and not prod: errors.append("產品代碼不存在")
        if _norm(row.get("診所代碼")) and not clinic: errors.append("診所代碼不存在")
        try: d=date.fromisoformat(_norm(row.get("日期")))
        except: d=None; errors.append("日期格式需為 YYYY-MM-DD")
        try: qty=float(row.get("數量",0)); assert qty>=0
        except: qty=0; errors.append("數量格式錯誤")
        try: amount=float(row.get("金額",0)); assert amount>=0
        except: amount=0; errors.append("金額格式錯誤")
        clean={"日期":_norm(row.get("日期")),"員工編號":_norm(row.get("員工編號")),"產品代碼":_norm(row.get("產品代碼")),"診所代碼":_norm(row.get("診所代碼")),"數量":qty,"金額":amount,"備註":_norm(row.get("備註"))}
        checked.append({"line":idx,"data":clean,"errors":errors})
    return checked

def _clinics_validate(rows, db):
    checked=[]; seen=set()
    for idx,row in enumerate(rows,start=2):
        code=_norm(row.get("客戶代碼")); name=_norm(row.get("診所名稱")); region=_norm(row.get("區域")); city=_norm(row.get("城市")); eno=_norm(row.get("負責業務員工編號")); status=_norm(row.get("狀態")) or "有效客戶"
        errors=[]
        if not code: errors.append("缺少客戶代碼")
        if not name: errors.append("缺少診所名稱")
        if region not in ["北區","中區","南區"]: errors.append("區域需為北區/中區/南區")
        if not city: errors.append("缺少城市")
        emp=db.scalar(select(Employee).where(Employee.employee_no==eno)) if eno else None
        if eno and not emp: errors.append("負責業務員工編號不存在")
        if not eno: errors.append("缺少負責業務員工編號")
        if status not in ["有效客戶","潛在客戶","暫停交易"]: errors.append("狀態不正確")
        if code and (db.scalar(select(Clinic).where(Clinic.code==code)) or code in seen): errors.append("客戶代碼重複")
        seen.add(code)
        checked.append({"line":idx,"data":{"客戶代碼":code,"診所名稱":name,"區域":region,"城市":city,"負責業務員工編號":eno,"狀態":status},"errors":errors})
    return checked

@app.get("/templates/sales-import.csv")
def sales_template(user:User=Depends(current_user)):
    text="日期,員工編號,產品代碼,診所代碼,數量,金額,備註\n2026-08-01,S001,METEORA,C001,2,800000,範例資料\n"
    return Response(content=text.encode("utf-8-sig"),media_type="text/csv",headers={"Content-Disposition":"attachment; filename=diamond_sales_import_template.csv"})

@app.get("/templates/clinics-import.csv")
def clinics_template(user:User=Depends(current_user)):
    text="客戶代碼,診所名稱,區域,城市,負責業務員工編號,狀態\nC001,範例醫美診所,北區,台北市,S001,有效客戶\n"
    return Response(content=text.encode("utf-8-sig"),media_type="text/csv",headers={"Content-Disposition":"attachment; filename=diamond_clinics_import_template.csv"})

@app.post("/import/sales/preview",response_class=HTMLResponse)
async def import_sales_preview(request:Request,file:UploadFile=File(...),db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin","executive","manager")
    try: rows=_read_tabular_upload(file.filename,await file.read())
    except Exception as e: raise HTTPException(400,f"無法讀取檔案：{e}")
    checked=_sales_validate(rows,db); token=signer.dumps({"kind":"sales","rows":[x["data"] for x in checked if not x["errors"]]})
    return templates.TemplateResponse("import_preview.html",{"request":request,"user":user,"company":COMPANY_NAME,"kind":"業績","checked":checked,"token":token,"confirm_url":"/import/sales/confirm"})

@app.post("/import/sales/confirm")
def import_sales_confirm(token:str=Form(...),db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin","executive","manager")
    try: payload=signer.loads(token)
    except BadSignature: raise HTTPException(400,"匯入資料已失效")
    if payload.get("kind")!="sales": raise HTTPException(400,"匯入類型錯誤")
    count=0
    for row in payload["rows"]:
        emp=db.scalar(select(Employee).where(Employee.employee_no==row["員工編號"])); prod=db.scalar(select(Product).where(Product.code==row["產品代碼"])); clinic=db.scalar(select(Clinic).where(Clinic.code==row["診所代碼"]))
        if not(emp and prod and clinic): continue
        amount=float(row["金額"]); db.add(Sale(sale_date=date.fromisoformat(row["日期"]),employee_id=emp.id,product_id=prod.id,clinic_id=clinic.id,quantity=float(row["數量"]),amount=amount,gross_profit=amount*prod.gross_margin,note=row.get("備註",""))); count+=1
    db.commit(); audit(db,user,"匯入","業績",f"{count} 筆（智慧匯入）")
    return RedirectResponse("/sales",303)

@app.post("/import/clinics/preview",response_class=HTMLResponse)
async def import_clinics_preview(request:Request,file:UploadFile=File(...),db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin","executive","manager")
    try: rows=_read_tabular_upload(file.filename,await file.read())
    except Exception as e: raise HTTPException(400,f"無法讀取檔案：{e}")
    checked=_clinics_validate(rows,db); token=signer.dumps({"kind":"clinics","rows":[x["data"] for x in checked if not x["errors"]]})
    return templates.TemplateResponse("import_preview.html",{"request":request,"user":user,"company":COMPANY_NAME,"kind":"診所","checked":checked,"token":token,"confirm_url":"/import/clinics/confirm"})

@app.post("/import/clinics/confirm")
def import_clinics_confirm(token:str=Form(...),db:Session=Depends(db_session),user:User=Depends(current_user)):
    authorize(user,"admin","executive","manager")
    try: payload=signer.loads(token)
    except BadSignature: raise HTTPException(400,"匯入資料已失效")
    if payload.get("kind")!="clinics": raise HTTPException(400,"匯入類型錯誤")
    count=0
    for row in payload["rows"]:
        if db.scalar(select(Clinic).where(Clinic.code==row["客戶代碼"])): continue
        emp=db.scalar(select(Employee).where(Employee.employee_no==row["負責業務員工編號"]))
        if not emp: continue
        db.add(Clinic(code=row["客戶代碼"],name=row["診所名稱"],region=row["區域"],city=row["城市"],owner_employee_id=emp.id,status=row["狀態"])); count+=1
    db.commit(); audit(db,user,"匯入","診所",f"{count} 筆（智慧匯入）")
    return RedirectResponse("/clinics",303)

@app.get("/health")
def health(): return {"status":"ok","time":datetime.utcnow().isoformat()}
